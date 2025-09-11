import os
import shutil
from datetime import datetime, timedelta
from pathlib import Path
from typing import List
import logging

from fastapi import APIRouter, File, UploadFile, HTTPException, BackgroundTasks, Request
from fastapi.responses import FileResponse, Response
import pandas as pd
import tabula
import fitz  # PyMuPDF

from core.config import settings
from core.utils import validate_pdf_file, save_upload_file, cleanup_old_files, ensure_safe_path

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/tools/pdf-to-excel", tags=["pdf-to-excel"])


@router.post("/upload")
async def upload_pdf_for_excel(
    background_tasks: BackgroundTasks,
    files: List[UploadFile] = File(...),
):
    if len(files) < 1:
        raise HTTPException(status_code=400, detail="En az 1 PDF dosyası gereklidir")
    if len(files) > settings.MAX_FILES:
        raise HTTPException(status_code=400, detail=f"Maksimum {settings.MAX_FILES} dosya yüklenebilir")

    session_id = datetime.now().strftime("%Y%m%d_%H%M%S_") + os.urandom(4).hex()
    session_dir = os.path.join(settings.TEMP_DIR, session_id)
    os.makedirs(session_dir, exist_ok=True)

    uploaded_files = []
    total_size = 0

    try:
        for idx, file in enumerate(files):
            validate_pdf_file(file)
            if getattr(file, "size", None) is not None:
                total_size += file.size
                if total_size > settings.MAX_FILE_SIZE:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Toplam dosya boyutu {settings.MAX_FILE_SIZE/(1024*1024)}MB sınırını aşıyor",
                    )

            file_path = Path(session_dir) / f"{idx}_{file.filename}"
            await save_upload_file(file, file_path)

            uploaded_files.append({
                "original_name": file.filename,
                "path": str(file_path),
                "size": getattr(file, "size", None) or 0,
            })
            logger.info(f"Dosya yüklendi: {file.filename}")

        if background_tasks:
            background_tasks.add_task(cleanup_old_files)

        return {
            "session_id": session_id,
            "files": uploaded_files,
            "total_files": len(uploaded_files),
            "total_size_mb": round((total_size or 0) / (1024 * 1024), 2),
        }

    except Exception as e:
        if os.path.exists(session_dir):
            shutil.rmtree(session_dir)
        if isinstance(e, HTTPException):
            raise e
        logger.error(f"Dosya yükleme hatası: {str(e)}")
        raise HTTPException(status_code=500, detail="Dosya yükleme sırasında hata oluştu")


@router.post("/process/{session_id}")
async def process_pdf_to_excel(
    session_id: str,
    background_tasks: BackgroundTasks,
    pages: str = "all",
    password: str = None,
    merge_paragraphs: bool = True,
    min_paragraph_lines: int = 5,
    detection_method: str = "auto",
):
    session_dir = os.path.join(settings.TEMP_DIR, session_id)
    if not os.path.exists(session_dir):
        raise HTTPException(status_code=404, detail="Oturum bulunamadı veya süresi dolmuş")

    try:
        files = [str(f) for f in Path(session_dir).glob("*.pdf")]
        if len(files) < 1:
            raise HTTPException(status_code=400, detail="PDF dosyası bulunamadı")

        # İlk PDF dosyasını işle
        pdf_file = files[0]
        
        # Akıllı PDF analizi - tablo/paragraf ayrımı
        try:
            # Önce PDF'i analiz et - tablo var mı, paragraf var mı?
            page_analysis = _analyze_pdf_content(pdf_file, pages, password)
            
            # Analiz sonucuna göre işle
            if page_analysis['has_tables'] and page_analysis['has_paragraphs']:
                # Hibrit sayfa - hem tablo hem paragraf var
                tables = _process_hybrid_content(pdf_file, pages, password, detection_method, merge_paragraphs, min_paragraph_lines)
            elif page_analysis['has_tables']:
                # Sadece tablo var - tabula kullan
                tables = _extract_tables_only(pdf_file, pages, password, detection_method)
            else:
                # Sadece paragraf var - metin olarak çıkar
                tables = _extract_paragraphs_only(pdf_file, pages, password, merge_paragraphs, min_paragraph_lines)
                
        except Exception as e:
            logger.error(f"PDF analiz hatası: {str(e)}")
            # Fallback: eski yöntem
            try:
                # Önce tabula-py ile tabloları çıkar
                tables = []
                
                # Algılama metoduna göre tabula kullan
                try:
                    if detection_method == "lattice":
                        if pages == "all":
                            tables = tabula.read_pdf(
                                pdf_file, 
                                pages='all', 
                                multiple_tables=True, 
                                password=password,
                                lattice=True,
                                stream=False,
                                pandas_options={'header': None},
                                guess=True,
                                area=None,
                                relative_area=False
                            )
                        else:
                            tables = tabula.read_pdf(
                                pdf_file, 
                                pages=pages, 
                                multiple_tables=True, 
                                password=password,
                                lattice=True,
                                stream=False,
                                pandas_options={'header': None},
                                guess=True,
                                area=None,
                                relative_area=False
                            )
                    elif detection_method == "stream":
                        if pages == "all":
                            tables = tabula.read_pdf(
                                pdf_file, 
                                pages='all', 
                                multiple_tables=True, 
                                password=password,
                                stream=True,
                                lattice=False,
                                pandas_options={'header': None},
                                guess=True,
                                area=None,
                                relative_area=False
                            )
                        else:
                            tables = tabula.read_pdf(
                                pdf_file, 
                                pages=pages, 
                                multiple_tables=True, 
                                password=password,
                                stream=True,
                                lattice=False,
                                pandas_options={'header': None},
                                guess=True,
                                area=None,
                                relative_area=False
                            )
                    else:  # auto
                        # Önce lattice metodunu dene (çizgili tablolar için)
                        try:
                            if pages == "all":
                                tables = tabula.read_pdf(
                                    pdf_file, 
                                    pages='all', 
                                    multiple_tables=True, 
                                    password=password,
                                    lattice=True,
                                    stream=False,
                                    pandas_options={'header': None},
                                    guess=True,
                                    area=None,
                                    relative_area=False
                                )
                            else:
                                tables = tabula.read_pdf(
                                    pdf_file, 
                                    pages=pages, 
                                    multiple_tables=True, 
                                    password=password,
                                    lattice=True,
                                    stream=False,
                                    pandas_options={'header': None},
                                    guess=True,
                                    area=None,
                                    relative_area=False
                                )
                        except Exception as lattice_error:
                            logger.warning(f"Lattice metodu başarısız, stream metodunu deniyorum: {str(lattice_error)}")
                            
                            # Stream metodunu dene (çizgisiz tablolar için)
                            if pages == "all":
                                tables = tabula.read_pdf(
                                    pdf_file, 
                                    pages='all', 
                                    multiple_tables=True, 
                                    password=password,
                                    stream=True,
                                    lattice=False,
                                    pandas_options={'header': None},
                                    guess=True,
                                    area=None,
                                    relative_area=False
                                )
                            else:
                                tables = tabula.read_pdf(
                                    pdf_file, 
                                    pages=pages, 
                                    multiple_tables=True, 
                                    password=password,
                                    stream=True,
                                    lattice=False,
                                    pandas_options={'header': None},
                                    guess=True,
                                    area=None,
                                    relative_area=False
                                )
                except Exception as tabula_error:
                    logger.warning(f"Tabula başarısız: {str(tabula_error)}")
                    tables = []
                
                # Eğer tabula başarısız olursa, PyMuPDF ile metin çıkar
                if not tables or len(tables) == 0:
                    logger.info("Tabula başarısız, PyMuPDF ile metin çıkarılıyor...")
                    tables = _extract_text_with_pymupdf(pdf_file, pages, password)
                
            except Exception as e:
                logger.error(f"PDF işleme hatası: {str(e)}")
                raise HTTPException(status_code=400, detail="PDF'den veri çıkarılamadı. Dosya şifreli olabilir veya işlenemiyor.")

        if not tables or len(tables) == 0:
            raise HTTPException(status_code=400, detail="PDF dosyasında tablo bulunamadı")

        # Paragraf birleştirme özelliği aktifse tabloları işle
        if merge_paragraphs:
            tables = _merge_paragraphs_in_tables(tables, min_paragraph_lines)

        # Excel dosyası oluştur
        output_filename = f"excel_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = os.path.join(session_dir, output_filename)

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for i, table in enumerate(tables):
                if table is not None and not table.empty:
                    # Tablo adı
                    sheet_name = f"Tablo_{i+1}" if len(tables) > 1 else "Tablo"
                    
                    # Türkçe karakterleri koru ve temizle
                    table = table.fillna('')  # NaN değerleri boş string yap
                    
                    # Excel'e yaz
                    table.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Excel formatlamasını uygula
                    worksheet = writer.sheets[sheet_name]
                    _format_excel_worksheet(worksheet, table)

        # Dosya bilgilerini al
        file_size = os.path.getsize(output_path)
        file_size_mb = round(file_size / (1024 * 1024), 2)
        
        logger.info(f"PDF'den Excel dönüştürme başarılı: {session_id}")

        return {
            "success": True,
            "session_id": session_id,
            "output_file": output_filename,
            "file_info": {
                "file_size_mb": file_size_mb,
                "table_count": len(tables),
                "sheet_count": len([t for t in tables if t is not None and not t.empty])
            },
            "download_url": f"/api/tools/pdf-to-excel/download/{session_id}/{output_filename}",
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"İşlem hatası: {str(e)}")
        raise HTTPException(status_code=500, detail="PDF'den Excel dönüştürme sırasında hata oluştu")


@router.api_route("/download/{session_id}/{filename}", methods=["GET", "HEAD"])
async def download_excel_file(session_id: str, filename: str, request: Request):
    session_dir = os.path.join(settings.TEMP_DIR, session_id)
    file_path = os.path.join(session_dir, filename)

    if not os.path.exists(session_dir):
        logger.warning(f"Session bulunamadı: {session_id}")
        raise HTTPException(
            status_code=404,
            detail=f"İndirme oturumu bulunamadı veya süresi dolmuş ({settings.SESSION_LIFETIME_MINUTES} dakika). Dosyaları tekrar yükleyip işleyin.",
        )

    if not os.path.exists(file_path) or not filename.endswith('.xlsx'):
        logger.warning(f"Dosya bulunamadı: {file_path}")
        raise HTTPException(status_code=404, detail="Dosya bulunamadı veya silinmiş")

    ensure_safe_path(file_path, settings.TEMP_DIR)

    try:
        session_time = datetime.fromtimestamp(os.path.getctime(session_dir))
        if datetime.now() - session_time > timedelta(minutes=settings.SESSION_LIFETIME_MINUTES):
            logger.info(f"Session süresi dolmuş: {session_id}")
            raise HTTPException(
                status_code=410,
                detail=f"İndirme linki süresi dolmuş ({settings.SESSION_LIFETIME_MINUTES} dakika). Lütfen dosyaları tekrar işleyin ve daha hızlı indirin.",
            )
    except Exception as e:
        logger.error(f"Session time check failed: {e}")

    if request.method == "HEAD":
        logger.info(f"HEAD request: {filename} (Session: {session_id})")
        return Response(
            status_code=200,
            headers={
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "Content-Length": str(os.path.getsize(file_path)),
                "Content-Disposition": f"attachment; filename={filename}",
                "Cache-Control": "no-cache, no-store, must-revalidate",
                "Pragma": "no-cache",
                "Expires": "0",
            },
        )

    logger.info(f"Dosya indiriliyor: {filename} (Session: {session_id})")
    return FileResponse(
        path=file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache",
            "Expires": "0",
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "GET, POST, OPTIONS",
            "Access-Control-Allow-Headers": "Content-Type, Authorization",
            "Access-Control-Expose-Headers": "Content-Disposition, Content-Length, Content-Type"
        },
    )


def _format_excel_worksheet(worksheet, table):
    """Excel worksheet'ini formatla"""
    from openpyxl.styles import Alignment, Border, Side, Font
    from openpyxl.utils import get_column_letter
    
    # Hücre hizalaması ve metin kaydırma
    alignment = Alignment(
        wrap_text=True,
        vertical='top',
        horizontal='left'
    )
    
    # Kenarlık stili
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Font stili
    font = Font(name='Arial', size=10)
    
    # Tüm hücrelere format uygula
    for row in range(1, len(table) + 1):
        for col in range(1, len(table.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.alignment = alignment
            cell.border = thin_border
            cell.font = font
    
    # Sütun genişliklerini otomatik ayarla
    for col in range(1, len(table.columns) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        
        # Bu sütundaki en uzun metni bul
        for row in range(1, len(table) + 1):
            cell_value = str(worksheet.cell(row=row, column=col).value or '')
            if len(cell_value) > max_length:
                max_length = len(cell_value)
        
        # Sütun genişliğini ayarla (min 10, max 50)
        adjusted_width = min(max(max_length + 2, 10), 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Satır yüksekliklerini ayarla
    for row in range(1, len(table) + 1):
        worksheet.row_dimensions[row].height = 20


def _extract_text_with_pymupdf(pdf_file: str, pages: str, password: str = None) -> List[pd.DataFrame]:
    """PyMuPDF ile PDF'den metin çıkar ve tablo formatında döndür"""
    try:
        # PDF'i aç
        if password:
            pdf_document = fitz.open(pdf_file, password=password)
        else:
            pdf_document = fitz.open(pdf_file)
        
        # Sayfa aralığını belirle
        if pages == "all":
            page_range = range(len(pdf_document))
        else:
            page_range = [int(pages) - 1]  # 0-indexed
        
        tables = []
        
        for page_num in page_range:
            page = pdf_document[page_num]
            
            # Sayfadaki metin bloklarını al
            text_blocks = page.get_text("dict")
            
            # Metin bloklarını işle
            rows = []
            current_row = []
            
            for block in text_blocks["blocks"]:
                if "lines" in block:  # Metin bloğu
                    for line in block["lines"]:
                        line_text = ""
                        for span in line["spans"]:
                            line_text += span["text"]
                        
                        if line_text.strip():
                            # Satır sonlarını temizle ve hücrelere böl
                            cells = line_text.split('\t') if '\t' in line_text else [line_text]
                            current_row.extend(cells)
                            
                            # Eğer satır tamamlandıysa (boş satır veya yeni paragraf)
                            if line_text.strip() == "" or len(current_row) > 5:  # Maksimum 5 hücre
                                if current_row:
                                    rows.append(current_row)
                                    current_row = []
            
            # Son satırı ekle
            if current_row:
                rows.append(current_row)
            
            # DataFrame oluştur
            if rows:
                # Maksimum sütun sayısını bul
                max_cols = max(len(row) for row in rows) if rows else 1
                
                # Tüm satırları aynı sütun sayısına getir
                normalized_rows = []
                for row in rows:
                    while len(row) < max_cols:
                        row.append("")
                    normalized_rows.append(row[:max_cols])  # Fazla sütunları kes
                
                # DataFrame oluştur
                df = pd.DataFrame(normalized_rows)
                tables.append(df)
        
        pdf_document.close()
        return tables
        
    except Exception as e:
        logger.error(f"PyMuPDF metin çıkarma hatası: {str(e)}")
        return []


def _merge_paragraphs_in_tables(tables: List[pd.DataFrame], min_lines: int = 5) -> List[pd.DataFrame]:
    """Tablolardaki büyük paragrafları tek hücrede birleştir"""
    processed_tables = []
    
    for table in tables:
        if table is None or table.empty:
            processed_tables.append(table)
            continue
            
        # Tabloyu kopyala
        processed_table = table.copy()
        
        # Her satırı kontrol et
        merged_rows = []
        i = 0
        
        while i < len(processed_table):
            current_row = processed_table.iloc[i]
            
            # Bu satırın metin içeriğini kontrol et
            text_content = ' '.join([str(cell) for cell in current_row if pd.notna(cell) and str(cell).strip()])
            
            # Eğer bu satır tek hücrede uzun metin içeriyorsa ve minimum satır sayısından fazlaysa
            if len(text_content.split()) >= min_lines * 5:  # Yaklaşık kelime sayısı
                # Bu satırı tek hücrede birleştir
                merged_row = [''] * len(current_row)
                merged_row[0] = text_content
                merged_rows.append(merged_row)
                i += 1
            else:
                # Normal tablo satırı olarak ekle
                merged_rows.append(current_row.tolist())
                i += 1
        
        # Yeni DataFrame oluştur
        if merged_rows:
            new_df = pd.DataFrame(merged_rows, columns=processed_table.columns)
            processed_tables.append(new_df)
        else:
            processed_tables.append(processed_table)
    
    return processed_tables


def _analyze_pdf_content(pdf_file: str, pages: str, password: str = None) -> dict:
    """PDF içeriğini analiz et - tablo var mı, paragraf var mı?"""
    try:
        # PDF'i aç
        if password:
            pdf_document = fitz.open(pdf_file, password=password)
        else:
            pdf_document = fitz.open(pdf_file)
        
        # Sayfa aralığını belirle
        if pages == "all":
            page_range = range(len(pdf_document))
        else:
            page_range = [int(pages) - 1]
        
        has_tables = False
        has_paragraphs = False
        total_text_length = 0
        table_indicators = 0
        
        for page_num in page_range:
            page = pdf_document[page_num]
            
            # Sayfadaki metin bloklarını al
            text_blocks = page.get_text("dict")
            
            for block in text_blocks["blocks"]:
                if "lines" in block:  # Metin bloğu
                    for line in block["lines"]:
                        line_text = ""
                        for span in line["spans"]:
                            line_text += span["text"]
                        
                        if line_text.strip():
                            total_text_length += len(line_text)
                            
                            # Tablo göstergelerini kontrol et
                            if _is_table_line(line_text):
                                table_indicators += 1
                            else:
                                # Uzun metin paragraf göstergesi
                                if len(line_text.split()) > 10:
                                    has_paragraphs = True
        
        # Tablo varlığını belirle
        if table_indicators > 0 and total_text_length > 0:
            table_ratio = table_indicators / (total_text_length / 100)  # 100 karakter başına tablo göstergesi
            has_tables = table_ratio > 0.1  # %10'dan fazla tablo göstergesi varsa tablo var
        
        # Paragraf varlığını belirle
        if not has_paragraphs and total_text_length > 500:  # Uzun metin varsa paragraf var
            has_paragraphs = True
        
        pdf_document.close()
        
        return {
            'has_tables': has_tables,
            'has_paragraphs': has_paragraphs,
            'total_text_length': total_text_length,
            'table_indicators': table_indicators
        }
        
    except Exception as e:
        logger.error(f"PDF analiz hatası: {str(e)}")
        return {'has_tables': False, 'has_paragraphs': True, 'total_text_length': 0, 'table_indicators': 0}


def _is_table_line(text: str) -> bool:
    """Metin satırının tablo satırı olup olmadığını kontrol et"""
    # Tablo göstergeleri
    table_indicators = [
        '|',  # Dikey çizgi
        '\t',  # Tab karakteri
        '  ',  # Çoklu boşluk (hizalama)
        '___',  # Alt çizgi
        '---',  # Tire
        '==='  # Eşittir
    ]
    
    # Sayısal veri kontrolü (tarih, fiyat, vs.)
    import re
    if re.search(r'\d+[.,]\d+', text):  # Sayısal değerler
        return True
    
    # Tablo göstergelerini kontrol et
    for indicator in table_indicators:
        if indicator in text:
            return True
    
    # Çok kısa satırlar (tablo başlığı olabilir)
    if len(text.strip()) < 50 and len(text.split()) < 5:
        return True
    
    return False


def _extract_tables_only(pdf_file: str, pages: str, password: str, detection_method: str) -> List[pd.DataFrame]:
    """Sadece tabloları çıkar"""
    try:
        if detection_method == "lattice":
            if pages == "all":
                tables = tabula.read_pdf(pdf_file, pages='all', multiple_tables=True, password=password, lattice=True, stream=False, pandas_options={'header': None}, guess=True)
            else:
                tables = tabula.read_pdf(pdf_file, pages=pages, multiple_tables=True, password=password, lattice=True, stream=False, pandas_options={'header': None}, guess=True)
        elif detection_method == "stream":
            if pages == "all":
                tables = tabula.read_pdf(pdf_file, pages='all', multiple_tables=True, password=password, stream=True, lattice=False, pandas_options={'header': None}, guess=True)
            else:
                tables = tabula.read_pdf(pdf_file, pages=pages, multiple_tables=True, password=password, stream=True, lattice=False, pandas_options={'header': None}, guess=True)
        else:  # auto
            try:
                if pages == "all":
                    tables = tabula.read_pdf(pdf_file, pages='all', multiple_tables=True, password=password, lattice=True, stream=False, pandas_options={'header': None}, guess=True)
                else:
                    tables = tabula.read_pdf(pdf_file, pages=pages, multiple_tables=True, password=password, lattice=True, stream=False, pandas_options={'header': None}, guess=True)
            except:
                if pages == "all":
                    tables = tabula.read_pdf(pdf_file, pages='all', multiple_tables=True, password=password, stream=True, lattice=False, pandas_options={'header': None}, guess=True)
                else:
                    tables = tabula.read_pdf(pdf_file, pages=pages, multiple_tables=True, password=password, stream=True, lattice=False, pandas_options={'header': None}, guess=True)
        
        return [table for table in tables if table is not None and not table.empty]
        
    except Exception as e:
        logger.error(f"Tablo çıkarma hatası: {str(e)}")
        return []


def _extract_paragraphs_only(pdf_file: str, pages: str, password: str, merge_paragraphs: bool, min_paragraph_lines: int) -> List[pd.DataFrame]:
    """Sadece paragrafları metin olarak çıkar"""
    try:
        # PDF'i aç
        if password:
            pdf_document = fitz.open(pdf_file, password=password)
        else:
            pdf_document = fitz.open(pdf_file)
        
        # Sayfa aralığını belirle
        if pages == "all":
            page_range = range(len(pdf_document))
        else:
            page_range = [int(pages) - 1]
        
        all_paragraphs = []
        
        for page_num in page_range:
            page = pdf_document[page_num]
            
            # Sayfadaki tüm metni al
            page_text = page.get_text()
            
            if page_text.strip():
                # Paragrafları ayır
                paragraphs = page_text.split('\n\n')
                
                for paragraph in paragraphs:
                    paragraph = paragraph.strip()
                    if paragraph and len(paragraph.split()) >= min_paragraph_lines:
                        # Paragrafı tek hücrede birleştir
                        all_paragraphs.append([paragraph])
        
        pdf_document.close()
        
        # DataFrame oluştur
        if all_paragraphs:
            df = pd.DataFrame(all_paragraphs, columns=['Metin'])
            return [df]
        else:
            return []
        
    except Exception as e:
        logger.error(f"Paragraf çıkarma hatası: {str(e)}")
        return []


def _process_hybrid_content(pdf_file: str, pages: str, password: str, detection_method: str, merge_paragraphs: bool, min_paragraph_lines: int) -> List[pd.DataFrame]:
    """Hibrit içerik - hem tablo hem paragraf var"""
    try:
        # Önce tabloları çıkar
        tables = _extract_tables_only(pdf_file, pages, password, detection_method)
        
        # Sonra paragrafları çıkar
        paragraphs = _extract_paragraphs_only(pdf_file, pages, password, merge_paragraphs, min_paragraph_lines)
        
        # Birleştir
        all_content = tables + paragraphs
        return all_content
        
    except Exception as e:
        logger.error(f"Hibrit içerik işleme hatası: {str(e)}")
        return []
